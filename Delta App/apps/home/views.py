# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us

"""
from django.shortcuts import render
from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from .models import classification, Ansad_db

import openpyxl as op
import pandas as pd 


import joblib 
import sklearn



def all_data(request):
    class_list=classification.objects.all()
    return render(request,'../templates/home/ui-forms.html',{'class_list':class_list})

def importer(request):
    return render(request,'../templates/home/import.html',{})    

def page_index(request):
    data=classification.objects.all()
    data_count = data.count()
    context = {
        'segment': 'page_index',
        'class_data':data,
        'data_count':data_count
    }
    html_template = loader.get_template('../templates/home/index.html')
    return HttpResponse(html_template.render(context, request)) 

    

@login_required(login_url="/login/")
def index(request):
    data=classification.objects.all()
    data_count = data.count()
    context = {
        'segment': 'charts-morris',
        'class_data':data,
        'data_count':data_count
    }
    html_template = loader.get_template('home/charts-morris.html')
    return HttpResponse(html_template.render(context, request))


@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))

def import_data(request):
    if request.method == 'POST':
        imported_data= request.FILES['myfile']
        print(imported_data)
        wb= op.load_workbook(imported_data)

        sheets = wb.sheetnames
        print(sheets)

        mydata=wb[sheets[0]]
        
        for row in mydata.iter_rows() :
            obj=Ansad_db.objects.create(
                annee=row[0].value,
                lib_chap=row[1].value,
                Libellé_SSCHAP=row[2].value,
                C_PART=row[3].value, 
                C_ART=row[4].value,
                C_PGPH=row[5].value,
                DOTATION=row[6].value,
                INDISPONIBLE=row[7].value,
                Indisp=row[8].value,
                DISPONIBLE=row[9].value,
                LIBELLE=row[10].value, 
            )
            obj.save()
    data_liste=Ansad_db.objects.all()
    return render(request,"../templates/home/ui-tables.html",{'data_liste':data_liste})   




def classifier(request):

    M_branche=joblib.load('C:/Users/Lenovo/myenv/Delta_APP/apps/NB_branche.csv')
    M_epi_epac=joblib.load('C:/Users/Lenovo/myenv/Delta_APP/apps/NB_epi_epac.csv')
    M_operation=joblib.load('C:/Users/Lenovo/myenv/Delta_APP/apps/NB_operation.csv')
    M_secteur=joblib.load('C:/Users/Lenovo/myenv/Delta_APP/apps/NB_secteur.csv')
    M_si=joblib.load('C:/Users/Lenovo/myenv/Delta_APP/apps/NB_si.csv')

    data_liste=Ansad_db.objects.all()
    
    for w in data_liste:
        w1=w.lib_chap
        w2=w.LIBELLE
        cls_secteur =  M_secteur.predict([str(w1)])
        cls_branche =  M_branche.predict([str(w1)])
        cls_epi_epac =  M_epi_epac.predict([str(w1)])
        cls_si =  M_si.predict([str(w1)])
        cls_operation =  M_operation.predict([str(w2)])
        print(w1,cls_secteur)


        obj=classification.objects.create(
                lib_chap=w.lib_chap,
                LIBELLE=w.LIBELLE,
                Epi_Epac=cls_epi_epac[0],
                Branche=cls_branche[0],
                Secteur=cls_secteur[0],
                Si=cls_si[0],
                Operation=cls_operation[0]
            )
        obj.save

    print(cls_branche[0])
    class_list=classification.objects.all()
    return render(request,"../templates/home/ui-forms.html",{'class_list':class_list})   
